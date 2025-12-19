from openpyxl import load_workbook
import pandas as pd

# Reload Excel file
file_path = "/mnt/data/IOP_Mapas_MiDia_CalculosT&A_MiComision (1).xlsx"
wb = load_workbook(file_path)
ws = wb.active

markets = []
texts = []

for row in ws.iter_rows(min_row=2):
    market = row[0].value  # Column A
    text = row[8].value    # Column I
    if market:
        markets.append(market)
        texts.append(text if text else "")

df = pd.DataFrame({
    "Mercado": markets,
    "Texto": texts
})

calculations = {
    "Horas efectivas": ["hora", "horas"],
    "Horas extra / Overtime": ["extra", "overtime", "moretime"],
    "Festivos": ["festivo"],
    "Domingos": ["domingo"],
    "Nocturnidad": ["nocturn"],
    "Ausencias": ["ausencia", "absence"],
    "Retroactividad": ["retro"],
    "Prorrateo": ["prorrate"]
}

matrix = {"Mercado": df["Mercado"]}

for calc, keywords in calculations.items():
    matrix[calc] = df["Texto"].apply(
        lambda t: "Sí" if any(k.lower() in t.lower() for k in keywords) else "No"
    )

result_df = pd.DataFrame(matrix)

# Export to Excel
output_path = "/mnt/data/matriz_calculos_por_mercado.xlsx"
result_df.to_excel(output_path, index=False)

output_path
